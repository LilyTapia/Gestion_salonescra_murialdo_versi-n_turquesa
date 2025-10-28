from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import user_passes_test
from django.contrib.auth import logout, login
from django.contrib.auth.models import User
from django.urls import reverse
from django.views.decorators.http import require_POST
from .forms import ReservationForm, BlackoutForm, MaterialForm, InventoryForm, InventoryUpdateForm, CustomUserCreationForm, AdminUserCreationForm
from django.http import HttpResponse
from django.utils import timezone
from datetime import time, datetime, date, timedelta
from django.db import transaction
from django.db.models import Count, Sum, Q
from collections import defaultdict
from .models import Room, Material, RoomInventory, Reservation, ReservationItem, Blackout, Notification
from .services import release_overdue_reservations, build_registration_metadata, get_reserved_material_quantity
from .dateutils import max_reservation_date
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.units import inch
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import io
import calendar



MONTH_NAMES = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio', 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
WEEKDAY_NAMES = ['Lun', 'Mar', 'Mie', 'Jue', 'Vie']

BASE_DAY_BLOCKS = (
    ('08:00', '08:45'),
    ('08:45', '09:30'),
    ('09:50', '10:35'),
    ('10:35', '11:20'),
    ('11:35', '12:20'),
    ('12:20', '13:05'),
    ('13:05', '13:50'),
    ('13:50', '14:35'),
    ('14:35', '15:20'),
    ('15:20', '16:05'),
    ('16:05', '16:50'),
)
THURSDAY_BLOCKS = BASE_DAY_BLOCKS + (('17:00', '18:00'),)
FRIDAY_BLOCKS = BASE_DAY_BLOCKS[:6]

WEEKDAY_BLOCK_SCHEDULE = {
    calendar.MONDAY: BASE_DAY_BLOCKS,
    calendar.TUESDAY: BASE_DAY_BLOCKS,
    calendar.WEDNESDAY: BASE_DAY_BLOCKS,
    calendar.THURSDAY: THURSDAY_BLOCKS,
    calendar.FRIDAY: FRIDAY_BLOCKS,
}


def get_blocks_for_weekday(weekday_index):
    blocks = WEEKDAY_BLOCK_SCHEDULE.get(weekday_index, ())
    results = []
    for idx, (start_str, end_str) in enumerate(blocks, start=1):
        results.append({
            'index': idx,
            'label': f'Bloque {idx}',
            'start_str': start_str,
            'end_str': end_str,
            'start_time': time.fromisoformat(start_str),
            'end_time': time.fromisoformat(end_str),
        })
    return results


def _match_reservation_blackouts(room, date_value, start_time, end_time):
    """Return blackouts generated for a reservation slot."""
    start_dt = datetime.combine(date_value, start_time)
    end_dt = datetime.combine(date_value, end_time)
    return list(
        Blackout.objects.filter(
            room=room,
            start_datetime=start_dt,
            end_datetime=end_dt,
            reason__startswith="Reserva de"
        )
    )


def get_unread_notifications(user):
    if not user.is_authenticated:
        return []
    unread = list(Notification.objects.filter(user=user, read_at__isnull=True).order_by('-created_at'))
    if unread:
        Notification.objects.filter(id__in=[note.id for note in unread]).update(read_at=timezone.now())
    return unread


def is_library_admin(user):
    return user.is_authenticated and (user.is_staff or user.groups.filter(name='AdminBiblioteca').exists())


def index(request):
    release_overdue_reservations()
    # Redirect unauthenticated users to login
    if not request.user.is_authenticated:
        return redirect('login')

    # Redirect based on user type
    if request.user.is_staff or request.user.groups.filter(name='AdminBiblioteca').exists():
        # Admin users go to dashboard
        return redirect('admin_dashboard')
    else:
        # Teachers and other users go to reservations
        return redirect('reservation_list')


@user_passes_test(is_library_admin)
def admin_dashboard(request):
    """Home section for library administrators with quick actions."""
    display_name = (
        request.user.get_full_name()
        or request.user.first_name
        or request.user.username
    )

    quick_links = [
        {
            'title': 'Reservas',
            'description': 'Revisa y administra todas las reservas creadas.',
            'url': reverse('reservation_list'),
            'cta': 'Ir a reservas',
            'tag': 'calendar',
        },
        {
            'title': 'Bloqueos',
            'description': 'Configura feriados y bloqueos de agenda.',
            'url': reverse('blackout_list'),
            'cta': 'Ir a bloqueos',
            'tag': 'stopwatch',
        },
        {
            'title': 'Materiales',
            'description': 'Actualiza el catálogo disponible para préstamo.',
            'url': reverse('material_list'),
            'cta': 'Ir a materiales',
            'tag': 'materials',
        },
        {
            'title': 'Inventario',
            'description': 'Controla el stock por salón y disponibilidad.',
            'url': reverse('inventory_list'),
            'cta': 'Ir a inventario',
            'tag': 'inventory',
        },
        {
            'title': 'Usuarios',
            'description': 'Gestiona docentes y permisos de acceso.',
            'url': reverse('user_list'),
            'cta': 'Ir a usuarios',
            'tag': 'users',
        },
        {
            'title': 'Reportes',
            'description': 'Consulta métricas y exporta información.',
            'url': reverse('reports'),
            'cta': 'Ir a reportes',
            'tag': 'reports',
        },
    ]

    teacher_count = User.objects.filter(groups__name='Docentes').distinct().count()
    if teacher_count == 0:
        teacher_count = User.objects.filter(is_staff=False).count()

    stats = {
        'reservations': Reservation.objects.count(),
        'materials': Material.objects.count(),
        'teachers': teacher_count,
        'today_reservations': Reservation.objects.filter(date=timezone.localdate()).count(),
    }

    context = {
        'display_name': display_name,
        'quick_links': quick_links,
        'stats': stats,
    }
    return render(request, 'dashboard/admin_home.html', context)

@user_passes_test(lambda u: u.is_authenticated)
def reservation_create(request):
    release_overdue_reservations()
    materials = list(Material.objects.order_by('name'))
    material_values = {m.id: '' for m in materials}
    if request.method == "POST":
        form = ReservationForm(request.POST, user=request.user)
        items = []
        for m in materials:
            raw_value = (request.POST.get(f"qty_{m.id}", "") or "").strip()
            material_values[m.id] = raw_value
            try:
                q = int(raw_value or 0)
            except (TypeError, ValueError):
                messages.error(request, "Las cantidades de materiales deben ser números enteros.")
                return redirect('reservation_create')
            if q < 0:
                messages.error(request, "Las cantidades de materiales no pueden ser negativas.")
                return redirect('reservation_create')
            if q > 0:
                items.append((m, q))
        if form.is_valid():
            room = form.cleaned_data["room"]
            date = form.cleaned_data["date"]
            start = form.cleaned_data["start_time"]
            end = form.cleaned_data["end_time"]
            course = form.cleaned_data["course"]
            subject = form.cleaned_data["subject"]

            today = timezone.localdate()
            if date < today:
                messages.error(request, "La fecha de la reserva debe ser igual o posterior a hoy.")
                return redirect('reservation_create')

            max_allowed = max_reservation_date(today)
            if date > max_allowed:
                messages.error(request, "Las reservas solo se permiten hasta con 1 mes de anticipación.")
                return redirect('reservation_create')

            # Validaciones simples (choque de reservas)
            exists = Reservation.objects.filter(room=room, date=date, start_time__lt=end, end_time__gt=start).exists()
            if exists:
                messages.error(request, "El salón ya está ocupado en ese horario.")
                return redirect('reservation_create')

            # Validación horario laboral
            if not (time(8,0) <= start < time(18,0) and time(8,0) < end <= time(18,0)):
                messages.error(request, "Horario permitido: 08:00 a 18:00.")
                return redirect('reservation_create')

            # Blackouts
            from datetime import datetime as _dt
            def _join(d,t): return _dt.combine(d,t)
            start_dt = _join(date, start); end_dt = _join(date, end)
            blackout_exists = Blackout.objects.filter(
                room__isnull=True, start_datetime__lt=end_dt, end_datetime__gt=start_dt
            ).exists() or Blackout.objects.filter(room=room, start_datetime__lt=end_dt, end_datetime__gt=start_dt).exists()
            if blackout_exists:
                messages.error(request, "Existe un bloqueo de agenda en ese horario (feriado/reunión).")
                return redirect('reservation_create')

            with transaction.atomic():
                # Stock availability check (per material)
                for material, qty in items:
                    inventory = (
                        RoomInventory.objects.select_for_update()
                        .filter(room=room, material=material)
                        .first()
                    )
                    if not inventory:
                        messages.error(
                            request,
                            f"No hay inventario configurado para {material.name} en ese sal\u00f3n.",
                        )
                        return redirect('reservation_create')

                    reserved_overlap = get_reserved_material_quantity(
                        room=room,
                        material_id=material.id,
                        date=date,
                        start_time=start,
                        end_time=end,
                    )
                    if reserved_overlap + qty > inventory.quantity:
                        messages.error(
                            request,
                            "No hay stock suficiente de materiales para ese sal\u00f3n.",
                        )
                        return redirect('reservation_create')

                r = Reservation.objects.create(
                    room=room,
                    date=date,
                    start_time=start,
                    end_time=end,
                    course=course,
                    subject=subject,
                    user=request.user
                )
                for material, qty in items:
                    ReservationItem.objects.create(reservation=r, material=material, quantity=qty)

                # Create blackout for the reservation
                start_dt = _join(date, start)
                end_dt = _join(date, end)
                username = request.user.username
                Blackout.objects.create(
                    room=room,
                    start_datetime=start_dt,
                    end_datetime=end_dt,
                    reason=f"Reserva de {username}",
                    created_by=request.user
                )

            messages.success(request, "Reserva creada con éxito.")
            return redirect('index')
    else:
        form = ReservationForm(user=request.user)
    context = {
        'form': form,
        'material_inputs': [(m, material_values.get(m.id, '')) for m in materials],
        'form_title': 'Nueva reserva',
        'submit_label': 'Crear reserva',
        'cancel_url': reverse('reservation_list'),
        'is_edit': False,
    }
    return render(request, 'reservation_form.html', context)


@user_passes_test(lambda u: u.is_authenticated)
def reservation_update(request, pk):
    release_overdue_reservations()
    reservation = get_object_or_404(
        Reservation.objects.select_related('room', 'user', 'course', 'subject').prefetch_related('items__material'),
        pk=pk
    )
    is_admin_user = request.user.is_staff or request.user.groups.filter(name='AdminBiblioteca').exists()
    if reservation.user_id != request.user.id and not is_admin_user:
        messages.error(request, "No tienes permiso para editar esta reserva.")
        return redirect('reservation_list')

    materials = list(Material.objects.order_by('name'))
    material_values = {m.id: '' for m in materials}

    if request.method == "POST":
        form = ReservationForm(request.POST, user=request.user)
        items = []
        for m in materials:
            raw_value = (request.POST.get(f"qty_{m.id}", "") or "").strip()
            material_values[m.id] = raw_value
            try:
                q = int(raw_value or 0)
            except (TypeError, ValueError):
                messages.error(request, "Las cantidades de materiales deben ser números enteros.")
                return redirect('reservation_update', pk=pk)
            if q < 0:
                messages.error(request, "Las cantidades de materiales no pueden ser negativas.")
                return redirect('reservation_update', pk=pk)
            if q > 0:
                items.append((m, q))

        if form.is_valid():
            room = form.cleaned_data["room"]
            date_value = form.cleaned_data["date"]
            start = form.cleaned_data["start_time"]
            end = form.cleaned_data["end_time"]
            course = form.cleaned_data["course"]
            subject = form.cleaned_data["subject"]

            today = timezone.localdate()
            if date_value < today:
                messages.error(request, "La fecha de la reserva debe ser igual o posterior a hoy.")
                return redirect('reservation_update', pk=pk)

            max_allowed = max_reservation_date(today)
            if date_value > max_allowed:
                messages.error(request, "Las reservas solo se permiten hasta con 1 mes de anticipación.")
                return redirect('reservation_update', pk=pk)

            exists = (
                Reservation.objects
                .filter(room=room, date=date_value, start_time__lt=end, end_time__gt=start)
                .exclude(pk=reservation.pk)
                .exists()
            )
            if exists:
                messages.error(request, "El salón ya está ocupado en ese horario.")
                return redirect('reservation_update', pk=pk)

            if not (time(8,0) <= start < time(18,0) and time(8,0) < end <= time(18,0)):
                messages.error(request, "Horario permitido: 08:00 a 18:00.")
                return redirect('reservation_update', pk=pk)

            old_blackouts = _match_reservation_blackouts(reservation.room, reservation.date, reservation.start_time, reservation.end_time)
            old_blackout_ids = [b.id for b in old_blackouts]

            start_dt = datetime.combine(date_value, start)
            end_dt = datetime.combine(date_value, end)
            blackout_exists = (
                Blackout.objects
                .filter(start_datetime__lt=end_dt, end_datetime__gt=start_dt)
                .filter(Q(room__isnull=True) | Q(room=room))
                .exclude(id__in=old_blackout_ids)
                .exists()
            )
            if blackout_exists:
                messages.error(request, "Existe un bloqueo de agenda en ese horario (feriado/reunión).")
                return redirect('reservation_update', pk=pk)

            with transaction.atomic():
                for material, qty in items:
                    inventory = (
                        RoomInventory.objects.select_for_update()
                        .filter(room=room, material=material)
                        .first()
                    )
                    if not inventory:
                        messages.error(
                            request,
                            f"No hay inventario configurado para {material.name} en ese sal\u00f3n.",
                        )
                        return redirect('reservation_update', pk=pk)

                    reserved_overlap = get_reserved_material_quantity(
                        room=room,
                        material_id=material.id,
                        date=date_value,
                        start_time=start,
                        end_time=end,
                        exclude_reservation_id=reservation.id,
                    )
                    if reserved_overlap + qty > inventory.quantity:
                        messages.error(
                            request,
                            "No hay stock suficiente de materiales para ese sal\u00f3n.",
                        )
                        return redirect('reservation_update', pk=pk)

                reservation.room = room
                reservation.date = date_value
                reservation.start_time = start
                reservation.end_time = end
                reservation.course = course
                reservation.subject = subject
                reservation.save()

                existing_items = {item.material_id: item for item in reservation.items.select_related('material')}
                new_material_ids = set()
                for material, qty in items:
                    new_material_ids.add(material.id)
                    item = existing_items.get(material.id)
                    if item:
                        if item.quantity != qty:
                            item.quantity = qty
                            item.save(update_fields=['quantity'])
                    else:
                        ReservationItem.objects.create(reservation=reservation, material=material, quantity=qty)

                for material_id, item in existing_items.items():
                    if material_id not in new_material_ids:
                        item.delete()

                blackout_owner = reservation.user or request.user
                reason_username = reservation.user.username if reservation.user else request.user.username
                if old_blackouts:
                    for blackout in old_blackouts:
                        blackout.room = room
                        blackout.start_datetime = start_dt
                        blackout.end_datetime = end_dt
                        blackout.reason = f"Reserva de {reason_username}"
                        blackout.created_by = blackout_owner
                        blackout.save(update_fields=['room', 'start_datetime', 'end_datetime', 'reason', 'created_by'])
                else:
                    Blackout.objects.create(
                        room=room,
                        start_datetime=start_dt,
                        end_datetime=end_dt,
                        reason=f"Reserva de {reason_username}",
                        created_by=blackout_owner,
                    )

            messages.success(request, "Reserva actualizada con éxito.")
            return redirect('reservation_list')
    else:
        initial_data = {
            'room': reservation.room,
            'date': reservation.date,
            'start_time': reservation.start_time,
            'end_time': reservation.end_time,
            'course': reservation.course,
            'subject': reservation.subject,
        }
        form = ReservationForm(initial=initial_data, user=request.user)
        for item in reservation.items.all():
            material_values[item.material_id] = str(item.quantity)
    context = {
        'form': form,
        'reservation': reservation,
        'material_inputs': [(m, material_values.get(m.id, '')) for m in materials],
        'form_title': 'Editar reserva',
        'submit_label': 'Actualizar reserva',
        'cancel_url': reverse('reservation_list'),
        'is_edit': True,
    }
    return render(request, 'reservation_form.html', context)


@require_POST
@user_passes_test(lambda u: u.is_authenticated)
def reservation_cancel(request, pk):
    release_overdue_reservations()
    reservation = get_object_or_404(
        Reservation.objects.select_related('room', 'user').prefetch_related('items__material'),
        pk=pk
    )
    is_admin_user = request.user.is_staff or request.user.groups.filter(name='AdminBiblioteca').exists()
    if reservation.user_id != request.user.id and not is_admin_user:
        messages.error(request, "No tienes permiso para cancelar esta reserva.")
        return redirect('reservation_list')

    blackouts = _match_reservation_blackouts(reservation.room, reservation.date, reservation.start_time, reservation.end_time)
    blackout_ids = [b.id for b in blackouts]

    with transaction.atomic():
        reservation.release_inventory(items=reservation.items.select_related('material'))
        reservation.delete()
        if blackout_ids:
            Blackout.objects.filter(id__in=blackout_ids).delete()

    messages.success(request, "Reserva cancelada con éxito.")
    return redirect('reservation_list')


def reservation_list(request):
    """List reservations - teachers see only their own, admins see all"""
    release_overdue_reservations()
    notifications = []
    if request.user.is_authenticated:
        is_admin = request.user.is_staff or request.user.groups.filter(name='AdminBiblioteca').exists()
        if is_admin:
            reservations = Reservation.objects.select_related('room', 'user', 'course', 'subject').prefetch_related('items__material').order_by('date', 'start_time', 'room__code')
        else:
            reservations = Reservation.objects.filter(user=request.user).select_related('room', 'user', 'course', 'subject').prefetch_related('items__material').order_by('date', 'start_time', 'room__code')
        notifications = get_unread_notifications(request.user)
    else:
        reservations = Reservation.objects.none()
        is_admin = False

    context = {
        'reservations': reservations,
        'notifications': notifications,
        'active_view': 'history',
        'is_admin': is_admin,
    }
    return render(request, 'reservations/list.html', context)




def reservation_monthly(request):
    release_overdue_reservations()
    today = timezone.localdate()
    try:
        year = int(request.GET.get('year', today.year))
        month = int(request.GET.get('month', today.month))
        display_date = date(year, month, 1)
    except (TypeError, ValueError):
        display_date = date(today.year, today.month, 1)

    first_of_month = display_date
    first_workday = first_of_month
    while first_workday.month == display_date.month and first_workday.weekday() >= 5:
        first_workday += timedelta(days=1)
    if first_workday.month != display_date.month:
        first_workday = first_of_month

    if first_workday.weekday() == 0 and first_of_month.weekday() >= 5:
        start_date = first_workday
    else:
        start_date = first_workday - timedelta(days=first_workday.weekday())

    last_day = date(display_date.year, display_date.month, calendar.monthrange(display_date.year, display_date.month)[1])
    last_workday = last_day
    while last_workday.month == display_date.month and last_workday.weekday() >= 5:
        last_workday -= timedelta(days=1)
    if last_workday.month != display_date.month:
        last_workday = last_day

    if last_workday.weekday() < 4:
        end_date = last_workday + timedelta(days=(4 - last_workday.weekday()))
    else:
        end_date = last_workday

    reservations_qs = (
        Reservation.objects
        .select_related('room', 'user', 'course', 'subject')
        .filter(date__gte=start_date, date__lte=end_date)
        .order_by('date', 'start_time')
    )

    rooms = list(Room.objects.order_by('code'))
    reservations_by_day = defaultdict(lambda: defaultdict(list))
    for reservation in reservations_qs:
        reservations_by_day[reservation.date][reservation.room.code].append(reservation)

    tz = timezone.get_current_timezone()

    def to_local(dt):
        if timezone.is_naive(dt):
            return timezone.make_aware(dt, tz)
        return timezone.localtime(dt)

    blackout_range_start = datetime.combine(start_date, time.min)
    blackout_range_end = datetime.combine(end_date + timedelta(days=1), time.min)

    blackouts_qs = (
        Blackout.objects.select_related('room')
        .filter(start_datetime__lt=blackout_range_end, end_datetime__gt=blackout_range_start)
        .exclude(reason__startswith='Reserva de')
        .order_by('start_datetime')
    )

    variant_order = {'holiday': 0, 'general': 1, 'room': 2}
    blackouts_by_day = defaultdict(list)
    room_blackouts_by_day = defaultdict(lambda: defaultdict(list))
    for blackout in blackouts_qs:
        start_dt = to_local(blackout.start_datetime)
        end_dt = to_local(blackout.end_datetime)
        current = max(start_dt.date(), start_date)
        last = min(end_dt.date(), end_date)
        reason_text = (blackout.reason or '').strip()
        while current <= last:
            if current == start_dt.date() and current == end_dt.date():
                time_label = f"{start_dt:%H:%M} - {end_dt:%H:%M}"
            elif current == start_dt.date():
                time_label = f"Desde {start_dt:%H:%M}"
            elif current == end_dt.date():
                time_label = f"Hasta {end_dt:%H:%M}"
            else:
                time_label = "Todo el día"
            blackouts_by_day[current].append({
                'scope': blackout.display_scope,
                'type': blackout.display_type,
                'variant': blackout.style_variant,
                'reason': reason_text,
                'time_label': time_label,
            })
            if blackout.room_id:
                day_start_time = start_dt.time() if current == start_dt.date() else time.min
                day_end_time = end_dt.time() if current == end_dt.date() else time.max
                if day_end_time > day_start_time:
                    room_blackouts_by_day[current][blackout.room.code].append({
                        'scope': blackout.display_scope,
                        'reason': reason_text,
                        'time_label': time_label,
                        'start': day_start_time,
                        'end': day_end_time,
                    })
            current += timedelta(days=1)

    for day_key, items in blackouts_by_day.items():
        items.sort(key=lambda item: (variant_order.get(item['variant'], 99), item['scope']))
    for day_key, room_map in room_blackouts_by_day.items():
        for room_code, entries in room_map.items():
            entries.sort(key=lambda item: item['start'])

    weeks = []
    current_week_start = start_date
    while current_week_start <= end_date:
        week_days = []
        for day_offset in range(5):
            day = current_week_start + timedelta(days=day_offset)
            daily_map = reservations_by_day.get(day, {})
            day_blocks = get_blocks_for_weekday(day.weekday())
            room_blocks = []

            day_room_blackouts = room_blackouts_by_day.get(day, {})

            for room in rooms:
                room_reservations = sorted(
                    daily_map.get(room.code, []),
                    key=lambda r: r.start_time
                )
                room_blackouts = day_room_blackouts.get(room.code, [])

                def serialize_reservation(reservation):
                    if reservation.user:
                        full_name = (reservation.user.get_full_name() or '').strip()
                        teacher_name = full_name or reservation.user.username
                    else:
                        teacher_name = 'Sin usuario'
                    return {
                        'teacher': teacher_name,
                        'course': reservation.course.name if reservation.course else '',
                        'subject': reservation.subject.name if reservation.subject else '',
                        'time_range': f"{reservation.start_time.strftime('%H:%M')} - {reservation.end_time.strftime('%H:%M')}",
                    }

                formatted_entries = [serialize_reservation(res) for res in room_reservations]

                block_entries = []
                for block_def in day_blocks:
                    block_info = {
                        'index': block_def['index'],
                        'label': block_def['label'],
                        'time_label': f"{block_def['start_str']} - {block_def['end_str']}",
                        'status': 'available',
                    }
                    matching_blackout = next(
                        (
                            blk for blk in room_blackouts
                            if blk['start'] < block_def['end_time']
                            and blk['end'] > block_def['start_time']
                        ),
                        None
                    )
                    if matching_blackout:
                        block_info['status'] = 'blackout'
                        block_info['blackout'] = matching_blackout
                    else:
                        matching_reservation = next(
                            (
                                res for res in room_reservations
                                if res.start_time < block_def['end_time']
                                and res.end_time > block_def['start_time']
                            ),
                            None
                        )
                        if matching_reservation:
                            block_info['status'] = 'reserved'
                            block_info['reservation'] = serialize_reservation(matching_reservation)
                    block_entries.append(block_info)

                room_blocks.append({
                    'room': room,
                    'reservations': formatted_entries,
                    'blocks': block_entries,
                    'reserved_blocks': [item for item in block_entries if item['status'] == 'reserved'],
                    'available_blocks': [item for item in block_entries if item['status'] == 'available'],
                    'block_map': {item['index']: item for item in block_entries},
                    'has_block_schedule': bool(day_blocks),
                })
            day_blackouts = blackouts_by_day.get(day, [])

            full_block_blackouts = [item for item in day_blackouts if item['variant'] in ('holiday', 'general')]
            room_level_blackouts = [item for item in day_blackouts if item['variant'] == 'room']

            is_full_day_block = bool(full_block_blackouts)

            room_schedules_map = {rb['room'].code: {'room': rb['room'], 'blocks': [], 'has_content': False} for rb in room_blocks}
            block_schedule = []
            if room_blocks and day_blocks and not is_full_day_block:
                for block_def in day_blocks:
                    schedule_row = {
                        'index': block_def['index'],
                        'label': block_def['label'],
                        'time_label': f"{block_def['start_str']} - {block_def['end_str']}",
                        'rooms': [],
                    }
                    has_visible_room = False
                    for room_block in room_blocks:
                        block_info = room_block['block_map'].get(block_def['index'], {})
                        status = block_info.get('status', 'available')
                        schedule_row['rooms'].append({
                            'room_code': room_block['room'].code,
                            'status': status,
                            'reservation': block_info.get('reservation'),
                        })
                        room_entry = room_schedules_map[room_block['room'].code]
                        if status != 'blackout':
                            has_visible_room = True
                            room_entry['blocks'].append({
                                'label': block_def['label'],
                                'time_label': f"{block_def['start_str']} - {block_def['end_str']}",
                                'status': status,
                                'reservation': block_info.get('reservation'),
                            })
                            room_entry['has_content'] = True
                    if has_visible_room:
                        block_schedule.append(schedule_row)
            room_schedules = []
            if not is_full_day_block and room_schedules_map:
                for room in rooms:
                    entry = room_schedules_map.get(room.code)
                    if entry and entry['has_content']:
                        room_schedules.append({
                            'room': entry['room'],
                            'blocks': entry['blocks'],
                        })

            week_days.append({
                'date': day,
                'in_month': day.month == display_date.month,
                'is_today': day == today,
                'is_weekend': day.weekday() >= 5,
                'weekday_label': WEEKDAY_NAMES[day.weekday()],
                'room_blocks': [] if is_full_day_block else room_blocks,
                'blackouts': room_level_blackouts if not is_full_day_block else [],
                'full_block_blackouts': full_block_blackouts,
                'is_holiday': any(item['variant'] == 'holiday' for item in full_block_blackouts),
                'has_blackouts': bool(day_blackouts),
                'is_full_day_block': is_full_day_block,
                'block_schedule': block_schedule,
                'has_schedule': bool(block_schedule),
                'schedule_rooms': [rb['room'] for rb in room_blocks] if not is_full_day_block else [],
                'room_schedules': room_schedules,
            })

        weeks.append(week_days)
        current_week_start += timedelta(days=7)

    def shift_month(base, offset):
        month_value = base.month + offset
        year_value = base.year
        while month_value < 1:
            month_value += 12
            year_value -= 1
        while month_value > 12:
            month_value -= 12
            year_value += 1
        return date(year_value, month_value, 1)

    prev_date = shift_month(display_date, -1)
    next_date = shift_month(display_date, 1)

    month_label = f"{MONTH_NAMES[display_date.month]} {display_date.year}"

    is_admin = request.user.is_authenticated and (
        request.user.is_staff or request.user.groups.filter(name='AdminBiblioteca').exists()
    )

    notifications = []
    if request.user.is_authenticated and not is_admin:
        notifications = get_unread_notifications(request.user)

    context = {
        'weeks': weeks,
        'weekday_names': WEEKDAY_NAMES,
        'month_label': month_label,
        'current_month': display_date.month,
        'current_year': display_date.year,
        'prev_month': prev_date.month,
        'prev_year': prev_date.year,
        'next_month': next_date.month,
        'next_year': next_date.year,
        'today': today,
        'rooms': rooms,
        'notifications': notifications,
        'active_view': 'calendar',
        'is_admin': is_admin,
    }
    return render(request, 'reservations/calendar.html', context)

def blackout_list(request):
    # Only show administrative blackouts, not reservation-generated ones
    items = Blackout.objects.select_related('room').exclude(
        reason__startswith='Reserva de'
    ).order_by('-start_datetime')
    return render(request, 'blackouts/list.html', {'items': items})


def _cancel_overlapping_reservations(room, start_dt, end_dt, *, reason=None):
    """Cancel reservations that conflict with a blackout and restore inventory."""
    if room:
        overlapping = Reservation.objects.filter(
            room=room,
            date=start_dt.date(),
            start_time__lt=end_dt.time(),
            end_time__gt=start_dt.time()
        )
    else:
        overlapping = Reservation.objects.filter(
            date=start_dt.date(),
            start_time__lt=end_dt.time(),
            end_time__gt=start_dt.time()
        )

    cancelled_count = 0
    for reservation in overlapping:
        reservation.release_inventory(items=reservation.items.all())

        if reservation.user:
            reason_text = (reason.strip() or 'un bloqueo de agenda') if reason else 'un bloqueo de agenda'
            notification_message = (
                f"Tu reserva del salon {reservation.room.code} para el {reservation.date:%d/%m/%Y} "
                f"entre {reservation.start_time.strftime('%H:%M')} y {reservation.end_time.strftime('%H:%M')} fue cancelada debido a {reason_text}."
            )
            Notification.objects.create(user=reservation.user, message=notification_message)

        Blackout.objects.filter(
            room=reservation.room,
            reason=f"Reserva de {reservation.user.username}",
            start_datetime=datetime.combine(reservation.date, reservation.start_time),
            end_datetime=datetime.combine(reservation.date, reservation.end_time)
        ).delete()

        reservation.delete()
        cancelled_count += 1

    return cancelled_count


@user_passes_test(is_library_admin)
def blackout_create(request):
    if request.method == "POST":
        form = BlackoutForm(request.POST)
        if form.is_valid():
            occurrences = form.get_occurrences()
            if not occurrences:
                messages.error(request, "No se pudo determinar el horario del bloqueo.")
                return render(request, 'blackouts/form.html', {'form': form, 'title': 'Nuevo bloqueo'})

            room = form.cleaned_data.get('room')
            reason = form.cleaned_data.get('reason', '')

            total_cancelled = 0
            created_count = 0
            for start_dt, end_dt in occurrences:
                total_cancelled += _cancel_overlapping_reservations(room, start_dt, end_dt, reason=reason)

                obj = Blackout(
                    room=room,
                    reason=reason,
                    start_datetime=start_dt,
                    end_datetime=end_dt,
                    created_by=request.user
                )
                obj.save()
                created_count += 1

            if created_count > 1:
                base_msg = f"Se crearon {created_count} bloqueos."
            else:
                base_msg = "Bloqueo creado."

            if total_cancelled > 0:
                base_msg += f" Se cancelaron {total_cancelled} reserva(s) que se solapaban."

            messages.success(request, base_msg)
            return redirect('blackout_list')
    else:
        form = BlackoutForm()
    return render(request, 'blackouts/form.html', {'form': form, 'title': 'Nuevo bloqueo'})


@user_passes_test(is_library_admin)
def blackout_update(request, pk):
    obj = get_object_or_404(Blackout, pk=pk)
    if request.method == "POST":
        form = BlackoutForm(request.POST, instance=obj)
        if form.is_valid():
            occurrences = form.get_occurrences()
            if not occurrences:
                messages.error(request, "No se pudo determinar el horario del bloqueo.")
                return render(request, 'blackouts/form.html', {'form': form, 'title': 'Editar bloqueo'})

            start_dt, end_dt = occurrences[0]

            updated_obj = form.save(commit=False)
            updated_obj.start_datetime = start_dt
            updated_obj.end_datetime = end_dt

            cancelled_count = _cancel_overlapping_reservations(updated_obj.room, start_dt, end_dt, reason=updated_obj.reason)

            updated_obj.save()

            if cancelled_count > 0:
                messages.success(request, f"Bloqueo actualizado. Se cancelaron {cancelled_count} reserva(s) que se solapaban.")
            else:
                messages.success(request, "Bloqueo actualizado.")
            return redirect('blackout_list')
    else:
        form = BlackoutForm(instance=obj)
    return render(request, 'blackouts/form.html', {'form': form, 'title': 'Editar bloqueo'})

@user_passes_test(is_library_admin)
def blackout_delete(request, pk):
    obj = get_object_or_404(Blackout, pk=pk)
    if request.method == "POST":
        obj.delete()
        messages.success(request, "Bloqueo eliminado.")
        return redirect('blackout_list')
    return render(request, 'blackouts/confirm_delete.html', {'obj': obj})

# Material Management Views
@user_passes_test(is_library_admin)
def material_list(request):
    materials = Material.objects.order_by('name')
    return render(request, 'materials/list.html', {'materials': materials})

@user_passes_test(is_library_admin)
def material_create(request):
    if request.method == "POST":
        form = MaterialForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Material creado exitosamente.")
            return redirect('material_list')
    else:
        form = MaterialForm()
    return render(request, 'materials/form.html', {'form': form, 'title': 'Nuevo Material'})

@user_passes_test(is_library_admin)
def material_update(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == "POST":
        form = MaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, "Material actualizado exitosamente.")
            return redirect('material_list')
    else:
        form = MaterialForm(instance=material)
    return render(request, 'materials/form.html', {'form': form, 'title': 'Editar Material'})

@user_passes_test(is_library_admin)
def material_delete(request, pk):
    material = get_object_or_404(Material, pk=pk)
    if request.method == "POST":
        try:
            material.delete()
            messages.success(request, "Material eliminado exitosamente.")
        except Exception as e:
            messages.error(request, f"No se puede eliminar el material: {str(e)}")
        return redirect('material_list')
    return render(request, 'materials/delete.html', {'material': material})

# Inventory Management Views
@user_passes_test(is_library_admin)
def inventory_list(request):
    release_overdue_reservations()
    inventory = RoomInventory.objects.select_related('room', 'material').order_by('room__code', 'material__name')
    rooms = Room.objects.order_by('code')
    materials = Material.objects.order_by('name')
    today = timezone.localdate()
    selected_date_str = request.GET.get('date', today.isoformat())
    try:
        selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
    except ValueError:
        selected_date = today
        selected_date_str = today.isoformat()

    weekday_blocks = get_blocks_for_weekday(selected_date.weekday())
    block_param = request.GET.get('block')
    selected_block = None
    if weekday_blocks:
        if block_param:
            try:
                block_index = int(block_param)
                selected_block = next((b for b in weekday_blocks if b['index'] == block_index), None)
            except (ValueError, TypeError):
                selected_block = None
        if not selected_block:
            # Default to first block of the day
            selected_block = weekday_blocks[0]
    block_start = selected_block['start_time'] if selected_block else None
    block_end = selected_block['end_time'] if selected_block else None

    availability_options = []
    for block in weekday_blocks:
        availability_options.append({
            'index': block['index'],
            'label': f"{block['label']} ({block['start_str']} - {block['end_str']})"
        })

    # Annotate inventory with availability details for the selected block
    for item in inventory:
        reserved_quantity = 0
        if block_start and block_end:
            reserved_quantity = get_reserved_material_quantity(
                room=item.room,
                material_id=item.material_id,
                date=selected_date,
                start_time=block_start,
                end_time=block_end,
            )
        available_quantity = max(item.quantity - reserved_quantity, 0)
        if item.quantity <= 0:
            availability_status = 'empty'
        elif available_quantity <= 0:
            availability_status = 'empty'
        else:
            ratio = available_quantity / item.quantity if item.quantity else 0
            if ratio <= 0.5:
                availability_status = 'warning'
            else:
                availability_status = 'ok'
        item.selected_reserved_quantity = reserved_quantity
        item.selected_available_quantity = available_quantity
        item.selected_availability_status = availability_status

    return render(request, 'inventory/list.html', {
        'inventory': inventory,
        'rooms': rooms,
        'materials': materials,
        'selected_date': selected_date,
        'selected_date_str': selected_date_str,
        'selected_block': selected_block,
        'weekday_blocks': availability_options,
        'has_block_schedule': bool(selected_block),
    })

@user_passes_test(is_library_admin)
def inventory_create(request):
    if request.method == "POST":
        form = InventoryForm(request.POST)
        if form.is_valid():
            room = form.cleaned_data['room']
            material = form.cleaned_data['material']
            # Check if inventory already exists
            existing = RoomInventory.objects.filter(room=room, material=material).first()
            if existing:
                messages.error(request, f"Ya existe inventario para {material.name} en salón {room.code}")
                return render(request, 'inventory/form.html', {'form': form, 'title': 'Agregar Inventario'})
            form.save()
            messages.success(request, "Inventario agregado exitosamente.")
            return redirect('inventory_list')
    else:
        form = InventoryForm()
    return render(request, 'inventory/form.html', {'form': form, 'title': 'Agregar Inventario'})

@user_passes_test(is_library_admin)
def inventory_update(request, pk):
    inventory = get_object_or_404(RoomInventory, pk=pk)
    if request.method == "POST":
        form = InventoryUpdateForm(request.POST)
        if form.is_valid():
            action = form.cleaned_data['action']
            quantity = form.cleaned_data['quantity']
            
            if action == 'add':
                inventory.quantity += quantity
            elif action == 'remove':
                new_qty = inventory.quantity - quantity
                if new_qty < 0:
                    messages.error(request, "No se puede quitar más cantidad de la disponible.")
                    return render(request, 'inventory/update.html', {'form': form, 'inventory': inventory})
                inventory.quantity = new_qty
            elif action == 'set':
                inventory.quantity = quantity
            
            inventory.save()
            messages.success(request, f"Inventario actualizado: {inventory.material.name} en salón {inventory.room.code}")
            return redirect('inventory_list')
    else:
        form = InventoryUpdateForm()
    return render(request, 'inventory/update.html', {'form': form, 'inventory': inventory})

@user_passes_test(is_library_admin)
def inventory_delete(request, pk):
    inventory = get_object_or_404(RoomInventory, pk=pk)
    if request.method == "POST":
        inventory.delete()
        messages.success(request, "Inventario eliminado exitosamente.")
        return redirect('inventory_list')
    return render(request, 'inventory/delete.html', {'item': inventory})

def custom_logout(request):
    """Custom logout view that properly clears session and forces redirect"""
    logout(request)
    response = redirect('/')
    # Clear all cookies related to authentication
    response.delete_cookie('sessionid')
    response.delete_cookie('csrftoken')
    # Add cache control headers to prevent caching
    response['Cache-Control'] = 'no-cache, no-store, must-revalidate'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def user_register(request):
    """User registration view"""
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'Cuenta creada exitosamente para {user.username}!')
            return redirect('index')
    else:
        form = CustomUserCreationForm()

    metadata = build_registration_metadata()
    return render(request, 'registration/register.html', {
        'form': form,
        'registration_metadata': metadata,
    })


@user_passes_test(is_library_admin)
def user_list(request):
    """List all users - only accessible to admins"""
    users = User.objects.select_related().prefetch_related('groups').order_by('username')
    return render(request, 'users/list.html', {'users': users})


@user_passes_test(is_library_admin)
def user_create(request):
    """Create new user - only accessible to admins"""
    if request.method == 'POST':
        form = AdminUserCreationForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, f'Usuario {user.username} creado exitosamente.')
            return redirect('user_list')
    else:
        form = AdminUserCreationForm()

    metadata = build_registration_metadata()
    return render(request, 'users/form.html', {
        'form': form,
        'title': 'Nuevo Usuario',
        'registration_metadata': metadata,
    })


@user_passes_test(is_library_admin)
def _resolve_report_dates(request):
    """Parse and normalize report date filters."""
    start = request.GET.get('start_date')
    end = request.GET.get('end_date')
    today = date.today()
    had_error = False

    if not start or not end:
        return today.replace(day=1), today, had_error

    try:
        start_obj = datetime.strptime(start, '%Y-%m-%d').date()
        end_obj = datetime.strptime(end, '%Y-%m-%d').date()
    except ValueError:
        messages.error(request, "Formato de fecha inválido.")
        return today.replace(day=1), today, True

    if start_obj > end_obj:
        messages.error(request, "La fecha de inicio no puede ser posterior a la fecha de término. Se usará el rango corregido.")
        start_obj, end_obj = end_obj, start_obj
        had_error = True

    return start_obj, end_obj, had_error


def reports_view(request):
    """Reports view with date range and room filters"""
    release_overdue_reservations()
    # Get filter parameters
    room_filter = request.GET.get('room')

    start_date_obj, end_date_obj, _ = _resolve_report_dates(request)
    start_date = start_date_obj.strftime('%Y-%m-%d')
    end_date = end_date_obj.strftime('%Y-%m-%d')
    
    # Base queryset for reservations in date range
    reservations_qs = Reservation.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj
    )
    
    # Apply room filter if specified
    if room_filter:
        reservations_qs = reservations_qs.filter(room_id=room_filter)
    
    # Report 1: Reservations by room (count)
    room_stats = reservations_qs.values(
        'room__code'
    ).annotate(
        reservation_count=Count('id')
    ).order_by('room__code')
    
    # Report 2: Materials requested (sum by type)
    material_stats = ReservationItem.objects.filter(
        reservation__in=reservations_qs
    ).values(
        'material__name'
    ).annotate(
        total_quantity=Sum('quantity')
    ).order_by('material__name')
    
    # Get all rooms for filter dropdown
    rooms = Room.objects.order_by('code')
    
    context = {
        'start_date': start_date,
        'end_date': end_date,
        'room_filter': room_filter,
        'room_stats': room_stats,
        'material_stats': material_stats,
        'rooms': rooms,
        'total_reservations': reservations_qs.count(),
        'date_range_display': f"{start_date_obj.strftime('%d/%m/%Y')} - {end_date_obj.strftime('%d/%m/%Y')}"
    }
    
    return render(request, 'reports/dashboard.html', context)


@user_passes_test(is_library_admin)
def export_reports_pdf(request):
    """Export reports data to PDF"""
    # Get the same filter parameters as reports_view
    room_filter = request.GET.get('room')
    start_date_obj, end_date_obj, had_error = _resolve_report_dates(request)
    if had_error:
        return redirect('reports')
    start_date = start_date_obj.strftime('%Y-%m-%d')
    end_date = end_date_obj.strftime('%Y-%m-%d')
    # Get the same data as reports_view
    reservations_qs = Reservation.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj
    )
    
    if room_filter:
        reservations_qs = reservations_qs.filter(room_id=room_filter)
    
    room_stats = reservations_qs.values(
        'room__code'
    ).annotate(
        reservation_count=Count('id')
    ).order_by('room__code')
    
    material_stats = ReservationItem.objects.filter(
        reservation__in=reservations_qs
    ).values(
        'material__name'
    ).annotate(
        total_quantity=Sum('quantity')
    ).order_by('material__name')
    
    # Check if there's data to export
    if not reservations_qs.exists():
        messages.error(request, "No hay datos para exportar en el período seleccionado.")
        return redirect('reports')
    
    # Create PDF
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="reporte_biblioteca_{start_date}_{end_date}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=30,
        alignment=1  # Center alignment
    )
    
    # Title
    title = Paragraph("Reporte de Biblioteca", title_style)
    elements.append(title)
    
    # Date range
    date_range = Paragraph(
        f"Período: {start_date_obj.strftime('%d/%m/%Y')} - {end_date_obj.strftime('%d/%m/%Y')}",
        styles['Normal']
    )
    elements.append(date_range)
    elements.append(Spacer(1, 20))
    
    # Summary stats
    summary_data = [
        ['Métrica', 'Valor'],
        ['Total de reservas', str(reservations_qs.count())],
        ['Salones utilizados', str(len(room_stats))],
        ['Tipos de materiales', str(len(material_stats))]
    ]
    
    summary_table = Table(summary_data)
    summary_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    elements.append(summary_table)
    elements.append(Spacer(1, 30))
    
    # Room statistics
    elements.append(Paragraph("Reservas por Salón", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    if room_stats:
        room_data = [['Código de Salón', 'Cantidad de Reservas']]
        for stat in room_stats:
            room_data.append([
                f"Salón {stat['room__code']}",
                str(stat['reservation_count'])
            ])
        
        room_table = Table(room_data)
        room_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(room_table)
    else:
        elements.append(Paragraph("No hay datos de reservas para el período seleccionado.", styles['Normal']))
    
    elements.append(Spacer(1, 30))
    
    # Material statistics
    elements.append(Paragraph("Materiales Solicitados", styles['Heading2']))
    elements.append(Spacer(1, 12))
    
    if material_stats:
        material_data = [['Material', 'Cantidad Total Solicitada']]
        for stat in material_stats:
            material_data.append([
                stat['material__name'],
                str(stat['total_quantity'])
            ])
        
        material_table = Table(material_data)
        material_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(material_table)
    else:
        elements.append(Paragraph("No hay datos de materiales para el período seleccionado.", styles['Normal']))
    
    # Build PDF
    doc.build(elements)
    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    
    return response


@user_passes_test(is_library_admin)
def export_reports_excel(request):
    """Export reports data to Excel"""
    # Get the same filter parameters as reports_view
    room_filter = request.GET.get('room')
    start_date_obj, end_date_obj, had_error = _resolve_report_dates(request)
    if had_error:
        return redirect('reports')
    start_date = start_date_obj.strftime('%Y-%m-%d')
    end_date = end_date_obj.strftime('%Y-%m-%d')
    
    # Get the same data as reports_view
    reservations_qs = Reservation.objects.filter(
        date__gte=start_date_obj,
        date__lte=end_date_obj
    )
    
    if room_filter:
        reservations_qs = reservations_qs.filter(room_id=room_filter)
    
    room_stats = reservations_qs.values(
        'room__code'
    ).annotate(
        reservation_count=Count('id')
    ).order_by('room__code')
    
    material_stats = ReservationItem.objects.filter(
        reservation__in=reservations_qs
    ).values(
        'material__name'
    ).annotate(
        total_quantity=Sum('quantity')
    ).order_by('material__name')
    
    # Check if there's data to export
    if not reservations_qs.exists():
        messages.error(request, "No hay datos para exportar en el período seleccionado.")
        return redirect('reports')
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Create summary sheet
    summary_ws = wb.create_sheet("Resumen")
    
    # Header styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Title
    summary_ws['A1'] = "Reporte de Biblioteca"
    summary_ws['A1'].font = Font(bold=True, size=16)
    summary_ws.merge_cells('A1:C1')
    
    # Date range
    summary_ws['A3'] = f"Período: {start_date_obj.strftime('%d/%m/%Y')} - {end_date_obj.strftime('%d/%m/%Y')}"
    summary_ws.merge_cells('A3:C3')
    
    # Summary statistics
    summary_ws['A5'] = "Métrica"
    summary_ws['B5'] = "Valor"
    summary_ws['A5'].font = header_font
    summary_ws['A5'].fill = header_fill
    summary_ws['A5'].alignment = header_alignment
    summary_ws['B5'].font = header_font
    summary_ws['B5'].fill = header_fill
    summary_ws['B5'].alignment = header_alignment
    
    summary_ws['A6'] = "Total de reservas"
    summary_ws['B6'] = reservations_qs.count()
    summary_ws['A7'] = "Salones utilizados"
    summary_ws['B7'] = len(room_stats)
    summary_ws['A8'] = "Tipos de materiales"
    summary_ws['B8'] = len(material_stats)
    
    # Adjust column widths
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 15
    
    # Create room statistics sheet
    room_ws = wb.create_sheet("Reservas por Salón")
    
    # Headers
    room_ws['A1'] = "Código de Salón"
    room_ws['B1'] = "Cantidad de Reservas"
    
    for col in ['A1', 'B1']:
        room_ws[col].font = header_font
        room_ws[col].fill = header_fill
        room_ws[col].alignment = header_alignment
    
    # Data
    row = 2
    for stat in room_stats:
        room_ws[f'A{row}'] = f"Salón {stat['room__code']}"
        room_ws[f'B{row}'] = stat['reservation_count']
        row += 1
    
    # Adjust column widths
    room_ws.column_dimensions['A'].width = 20
    room_ws.column_dimensions['B'].width = 25
    
    # Create material statistics sheet
    material_ws = wb.create_sheet("Materiales Solicitados")
    
    # Headers
    material_ws['A1'] = "Material"
    material_ws['B1'] = "Cantidad Total Solicitada"
    
    for col in ['A1', 'B1']:
        material_ws[col].font = header_font
        material_ws[col].fill = header_fill
        material_ws[col].alignment = header_alignment
    
    # Data
    row = 2
    for stat in material_stats:
        material_ws[f'A{row}'] = stat['material__name']
        material_ws[f'B{row}'] = stat['total_quantity']
        row += 1
    
    # Adjust column widths
    material_ws.column_dimensions['A'].width = 30
    material_ws.column_dimensions['B'].width = 25
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_biblioteca_{start_date}_{end_date}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
